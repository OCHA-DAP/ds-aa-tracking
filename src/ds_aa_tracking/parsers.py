"""Parse the colleagues' AA tracking workbooks into tidy frames.

Each parse_* function returns {table_name: DataFrame}. Only *primary* data sheets are
parsed; derived pivot/scratch sheets in the same workbooks are deliberately skipped
(they recompute from the primary sheets). See docs on the review site for the
sheet-by-sheet decisions.

Raw workbooks are read from AA_TRACKING_DIR (default: ~/OCHA/data/aa_tracking) and are
never committed to this public repo.
"""

import os
import re
from pathlib import Path

import openpyxl
import pandas as pd

from .normalize import (
    iso3_from_application_code,
    norm_country,
    norm_fund,
    norm_hazard,
    norm_month,
    norm_status,
    subunit_of,
)

BASE = Path(os.environ.get("AA_TRACKING_DIR", "~/OCHA/data/aa_tracking")).expanduser()

F_PLANNING = BASE / "julia/2026 AA planning_USE THIS.xlsx"
F_REPORTING = BASE / "julia/AA reports_counting frameworks and countries.xlsx"
F_ACTIVATIONS = BASE / "julia/OCHA_AA_activations_2020-2026.xlsb"
F_ALLOC_ANALYSIS = BASE / "yakubu/26 March 2026 CERF Allocation Analysis for AA .xlsx"
F_DISPLACEMENT = BASE / "yakubu/AA Displacement Data - Clean.xlsx"
F_SUBGRANTS = BASE / "yakubu/CERF AA - Clean Subgrant  Data 2020 - 2025_AR.xlsx"
F_JUN2026 = BASE / "yakubu/CERF AA Data as of June 1st 2026.xlsx"

UNMAPPED: set = set()  # country names that failed iso3 resolution (reported at end)


def iso3(name):
    code, _ = norm_country(name)
    if code is None and name is not None and str(name).strip():
        try:
            import pycountry

            code = pycountry.countries.lookup(str(name).strip()).alpha_3
        except LookupError:
            UNMAPPED.add(str(name).strip())
    return code


def _s(v):
    """Cell → stripped string or None."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    return s or None


def _num(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if not isinstance(v, str) and pd.isna(v):
        return None
    try:
        s = str(v).replace("$", "").replace(",", "").strip()
        if s in {"-", "", "??", "TBC", "tbc"}:
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _bool(v, yes=("yes", "y", "true"), no=("no", "n", "false")):
    s = _s(v)
    if s is None:
        return None
    s = s.lower()
    if s in yes:
        return True
    if s in no:
        return False
    return None


def _strip_cols(df):
    df.columns = [str(c).strip() if not isinstance(c, int) else c for c in df.columns]
    return df


def _date(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    ts = pd.to_datetime(v, errors="coerce")
    return None if pd.isna(ts) else ts.date()


# --------------------------------------------------------------------------
# Julia: 2026 planning (registry, status, focal points, calendar, funding, Start)
# --------------------------------------------------------------------------

FILL_PHASE = {
    "FF00B050": "trigger_window",
    "FFFF0000": "proposal_development",
    "FFFFC000": "framework_finalization",
}
ROLES = {
    9: "support_lead", 10: "backup", 11: "chd", 12: "crd_desk",
    13: "cerf_fp", 14: "cbpf_fp", 15: "in_copy",
}


def parse_planning():
    src = "julia-planning-2026"
    as_of = "2026-08-01"
    wb = openpyxl.load_workbook(F_PLANNING)
    ws = wb["Status and planning"]

    registry, status, focal, calendar, funding, covered, start = [], [], [], [], [], [], []
    for r in range(14, ws.max_row + 1):
        country = _s(ws.cell(r, 3).value)
        hazard_raw = _s(ws.cell(r, 4).value)
        if not country or not hazard_raw:
            continue
        c = iso3(country)
        if c is None:
            continue
        h, hraw = norm_hazard(hazard_raw)
        stage, stage_raw = norm_status(ws.cell(r, 5).value)
        registry.append({
            "country_iso3": c, "hazard": h, "country_name": country, "hazard_raw": hraw,
            "region": _s(ws.cell(r, 2).value),
            "language": _s(ws.cell(r, 16).value),
            "us_prio": _s(ws.cell(r, 1).value) == "Y",
            "coordination_group": _s(ws.cell(r, 30).value),
        })
        if stage:
            status.append({
                "country_iso3": c, "hazard": h, "as_of": as_of, "source": src,
                "status": stage, "status_raw": stage_raw,
            })
        for col, role in ROLES.items():
            person = _s(ws.cell(r, col).value)
            if person:
                focal.append({
                    "country_iso3": c, "hazard": h, "role": role, "person": person,
                    "as_of": as_of, "source": src,
                })
        for col in range(17, 29):
            cell = ws.cell(r, col)
            fill = None
            if cell.fill is not None and cell.fill.patternType:
                fill = FILL_PHASE.get(cell.fill.fgColor.rgb)
            is_f = _s(cell.value) == "F"
            if fill or is_f:
                calendar.append({
                    "country_iso3": c, "hazard": h, "month": col - 16,
                    "phase": fill or "framework_finalization",
                    "is_finalization_deadline": is_f, "as_of": as_of, "source": src,
                })
        for col, fund in ((6, "cerf"), (7, "country_regional")):
            amt = _num(ws.cell(r, col).value)
            if amt is not None:
                funding.append({
                    "country_iso3": c, "hazard": h, "year": 2026,
                    "kind": "prearranged", "fund_source": fund, "amount_usd": amt,
                    "source": src,
                })
        pc = _num(ws.cell(r, 8).value)
        if pc is not None:
            covered.append({
                "country_iso3": c, "hazard": h, "as_of": as_of, "source": src,
                "people_covered": int(pc),
            })
        n_alerts = _num(ws.cell(r, 31).value)
        if n_alerts is not None or _s(ws.cell(r, 34).value):
            start.append({
                "country_iso3": c, "as_of": as_of, "source": src,
                "alerts_count": int(n_alerts) if n_alerts is not None else None,
                "alert_years": _s(ws.cell(r, 32).value),
                "alerts_activated": (
                    int(_num(ws.cell(r, 33).value))
                    if _num(ws.cell(r, 33).value) is not None else None
                ),
                "start_ready": _bool(ws.cell(r, 34).value),
            })
    start_df = pd.DataFrame(start).drop_duplicates(subset=["country_iso3"], keep="first")
    return {
        "framework_registry": pd.DataFrame(registry),
        "framework_status": pd.DataFrame(status),
        "framework_focal_point": pd.DataFrame(focal).drop_duplicates(),
        "framework_calendar": pd.DataFrame(calendar),
        "prearranged_funding": pd.DataFrame(funding),
        "people_covered": pd.DataFrame(covered),
        "start_network": start_df,
    }


# --------------------------------------------------------------------------
# Julia: reporting workbook (status per year, channels, funding, GHO)
# --------------------------------------------------------------------------

CH_2024 = [
    (3, "anticipation_hub", "framework"), (4, "uk_joint_humanitarian_bc", "framework"),
    (5, "uk_cerf_bc", "framework"), (6, "sg_natural_disasters", "framework"),
    (7, "cerf_annual_report", "framework"), (8, "cerf_annual_report", "country"),
    (9, "ocha_annual_report", "framework"), (10, "ocha_annual_report", "country"),
    (11, "ocha_sf_kpi", "framework"), (12, "cpc", "country"),
]
CH_2025 = [
    (15, "anticipation_hub", "framework"), (16, "uk_baseline", "framework"),
    (17, "uk_joint_humanitarian_bc", "framework"), (18, "sg_report", "framework"),
    (19, "cerf_annual_report", "framework"), (20, "cerf_annual_report", "country"),
    (21, "ocha_annual_report", "framework"), (22, "ocha_sf_kpi", "framework"),
    (23, "cpc", "country"),
]


def _channel_rows(df, year, channels, c, h, i, src):
    out = []
    for col, channel, unit in channels:
        if col >= df.shape[1]:
            continue
        v = _s(df.iat[i, col])
        if v is None:
            continue
        counted = v in {"1", "1.0"} or _num(v) == 1 or v.lower() in {
            "active", "under development", "yes",
        }
        out.append({
            "report_year": year, "channel": channel, "country_iso3": c, "hazard": h,
            "unit": unit, "counted": bool(counted),
            "note": v if not v.replace(".", "").isdigit() else None, "source": src,
        })
    return out


def parse_reporting():
    xl = pd.ExcelFile(F_REPORTING)
    status, channels, funding, covered, plans = [], [], [], [], []

    # ---- 2024 sheet
    df = xl.parse("2024 AA reporting", header=None)
    src = "julia-reporting-2024"
    for i in range(len(df)):
        c = iso3(_s(df.iat[i, 0])) if _s(df.iat[i, 0]) else None
        h, hraw = norm_hazard(_s(df.iat[i, 1]))
        if c is None or h is None:
            continue
        st, st_raw = norm_status(_s(df.iat[i, 2]))
        if st:
            status.append({
                "country_iso3": c, "hazard": h, "as_of": "2024-12-31", "source": src,
                "status": st, "status_raw": st_raw,
            })
        channels += _channel_rows(df, 2024, CH_2024, c, h, i, src)

    # ---- 2025 sheet
    df = xl.parse("2025 AA reporting", header=None)
    src = "julia-reporting-2025"
    for i in range(len(df)):
        c = iso3(_s(df.iat[i, 0])) if _s(df.iat[i, 0]) else None
        h, hraw = norm_hazard(_s(df.iat[i, 1]))
        if c is None or h is None or _s(df.iat[i, 0]) == "Countries":
            continue
        st, st_raw = norm_status(_s(df.iat[i, 2]))
        if st:
            status.append({
                "country_iso3": c, "hazard": h, "as_of": "2025-12-31", "source": src,
                "status": st, "status_raw": st_raw,
                "funding_change": _s(df.iat[i, 3]), "revised_on": _date(df.iat[i, 4]),
            })
        for year, cols in ((2025, (7, 8)), (2026, (10, 11))):
            for col, fund in zip(cols, ("cerf", "country_regional")):
                amt = _num(df.iat[i, col])
                if amt is not None:
                    funding.append({
                        "country_iso3": c, "hazard": h, "year": year,
                        "kind": "prearranged", "fund_source": fund,
                        "amount_usd": amt, "source": src,
                    })
        cof_flag = _bool(df.iat[i, 12])
        cof_amt = _num(df.iat[i, 13])
        if cof_flag is not None or cof_amt is not None:
            funding.append({
                "country_iso3": c, "hazard": h, "year": 2025, "kind": "cofinancing",
                "fund_source": "other", "amount_usd": cof_amt,
                "identified": cof_flag, "source": src,
            })
        pc = _num(df.iat[i, 14])
        if pc is not None:
            covered.append({
                "country_iso3": c, "hazard": h, "as_of": "2025-12-31", "source": src,
                "people_covered": int(pc),
            })
        gho = _s(df.iat[i, 5])
        if gho in {"Y", "N"}:
            plans.append({
                "country_iso3": c, "year": 2025, "source": src, "in_gho": gho == "Y",
            })
        channels += _channel_rows(df, 2025, CH_2025, c, h, i, src)

    # ---- 2026 GHO sheet
    df = _strip_cols(xl.parse("2026 GHO", header=2))
    src = "julia-gho-2026"
    for _, row in df.iterrows():
        c = iso3(_s(row.get("Countries")))
        h, hraw = norm_hazard(_s(row.get("Hazard")))
        if c is None or h is None:
            continue
        st, st_raw = norm_status(_s(row.get("Status")))
        if st:
            status.append({
                "country_iso3": c, "hazard": h, "as_of": "2025-12-01", "source": src,
                "status": st, "status_raw": st_raw,
                "revised_on": _date(row.get("Time of Revision")),
            })
        cols = {str(col).strip(): col for col in df.columns}
        amt25 = _num(row.get(cols.get("Pre-Arranged Funding Amount\n2025")))
        amt26 = _num(row.get(cols.get("Likely Pre-Arranged Funding Amount\n2026")))
        for year, amt in ((2025, amt25), (2026, amt26)):
            if amt is not None:
                funding.append({
                    "country_iso3": c, "hazard": h, "year": year, "kind": "prearranged",
                    "fund_source": "all", "amount_usd": amt, "source": src,
                })
        hnrp = _s(row.get("2026 GHO - HNRP"))
        if hnrp in {"Y", "N"}:
            plans.append({
                "country_iso3": c, "year": 2026, "source": src, "in_gho": hnrp == "Y",
                "plan_type": "HNRP" if hnrp == "Y" else None,
            })
    plans_df = pd.DataFrame(plans)
    if not plans_df.empty:
        plans_df = (
            plans_df.sort_values("in_gho", ascending=False)
            .drop_duplicates(subset=["country_iso3", "year", "source"], keep="first")
        )
    return {
        "framework_status": pd.DataFrame(status),
        "report_channel_inclusion": pd.DataFrame(channels).drop_duplicates(
            subset=["report_year", "channel", "country_iso3", "hazard", "unit"]
        ),
        "prearranged_funding": pd.DataFrame(funding),
        "people_covered": pd.DataFrame(covered),
        "plan_inclusion": plans_df,
    }


# --------------------------------------------------------------------------
# Julia: activations 2020-2026 (the superset activation-event list)
# --------------------------------------------------------------------------

def parse_activations():
    src = "julia-activations"
    df = _strip_cols(pd.read_excel(F_ACTIVATIONS, sheet_name="Overall activations",
                                   header=1, engine="pyxlsb"))
    rows = []
    for _, r in df.iterrows():
        country = _s(r.get("Country"))
        if not country or pd.isna(r.get("Year")):
            continue
        c = iso3(country)
        if c is None:
            continue
        h, hraw = norm_hazard(_s(r.get("Hazard")))
        mech = (_s(r.get("Framework?")) or "").lower()
        rows.append({
            "country_iso3": c, "hazard": h, "year": int(r["Year"]),
            "month": norm_month(_s(r.get("Month"))),
            "fund_source": norm_fund(_s(r.get("Fund"))),
            "mechanism": "framework" if mech.startswith("framework") else "adhoc",
            "aa_or_ea": _s(r.get("AA / EA")),
            "amount_usd": _num(r.get("Amount")),
            "people_targeted": (
                int(_num(r.get("People targeted")))
                if _num(r.get("People targeted")) is not None else None
            ),
            "reported_to_ahub": (_s(r.get("Reported to A-Hub")) or "").lower() or None,
            "region": _s(r.get("Region")),
            "comments": _s(r.get("Comments")),
            "source": src,
        })
    return {"activation_event": pd.DataFrame(rows)}


# --------------------------------------------------------------------------
# Yakubu: March 2026 allocation analysis (retags, HNRP flags, CIRV, GHO stats)
# --------------------------------------------------------------------------

def parse_alloc_analysis():
    xl = pd.ExcelFile(F_ALLOC_ANALYSIS)
    out = {}

    df = _strip_cols(xl.parse("Retagged Allocations"))
    rows = []
    for _, r in df.iterrows():
        code = _s(r.get("Code"))
        if not code:
            continue
        actual = _s(r.get("Actual shock")) or ""
        m = re.match(r"^(.*?)\s*\((.+)\)\s*$", actual)
        rows.append({
            "application_code": code,
            "country_iso3": iso3(_s(r.get("Country"))),
            "country_name": _s(r.get("Country")),
            "initial_type": _s(r.get("Initial shock in GMS")),
            "actual_type": m.group(1) if m else actual,
            "storm_name": m.group(2) if m else None,
            "amount_usd": _num(r.get("Amount")),
            "source": "yakubu-retagged-mar2026",
        })
    out["emergency_type_override"] = pd.DataFrame(rows)

    df = _strip_cols(xl.parse("CIRV"))
    rows = []
    for _, r in df.iterrows():
        c = iso3(_s(r.get("Country")))
        v = _num(r.get("CIRV"))
        if c and v is not None:
            rows.append({
                "country_iso3": c, "year": 2025, "country_name": _s(r.get("Country")),
                "cirv": v, "source": "yakubu-cirv-2025",
            })
    out["cirv"] = pd.DataFrame(rows).drop_duplicates(subset=["country_iso3", "year"])

    plans = []
    df = _strip_cols(xl.parse("HNRP 2025 countries"))
    for _, r in df.iterrows():
        c = iso3(_s(r.get("Country")))
        if c is None:
            continue
        plans.append({
            "country_iso3": c, "year": 2025, "source": "yakubu-hnrp-2025",
            "plan_type": _s(r.get("HPC/Plan")),
            "exposure_aa_shocks": _s(r.get("Exposure to AA-relevant shocks")),
            "aa_feasible": _bool(r.get("AA currently feasible?")),
            "aa_prearranged": _bool(r.get("AA $$ pre-arranged by CERF?")),
            "has_framework": _bool(r.get("Framework")),
        })

    df = _strip_cols(xl.parse("HNRP FCDO BC analysis"))
    for _, r in df.iterrows():
        c = iso3(_s(r.get("Country")))
        if c is None:
            continue
        for year, fcol, pcol in ((2025, "2025 Framework", "2025 HNRP"),
                                 (2026, "2026 Framework", "2026 HNRP")):
            plans.append({
                "country_iso3": c, "year": year, "source": "yakubu-fcdo-bc",
                "plan_type": _s(r.get(pcol)),
                "has_framework": _bool(r.get(fcol), yes=("yes",), no=("no",)),
            })

    df = xl.parse("Summary by country", header=None)
    for i in range(2, len(df)):
        c = iso3(_s(df.iat[i, 0]))
        if c is None:
            continue
        plans.append({
            "country_iso3": c, "year": 2025, "source": "yakubu-summary-mar2026",
            "plan_type": _s(df.iat[i, 14]),
            "aa_prearranged": _bool(df.iat[i, 4]),
            "gho_target_people": (
                int(_num(df.iat[i, 15])) if _num(df.iat[i, 15]) is not None else None
            ),
            "gho_requirement_usd": _num(df.iat[i, 16]),
        })
    out["plan_inclusion"] = pd.DataFrame(plans).drop_duplicates(
        subset=["country_iso3", "year", "source"]
    )
    return out


# --------------------------------------------------------------------------
# Yakubu: displacement workbook (GMS demographics + narratives, OneGMS extra)
# --------------------------------------------------------------------------

SEX_AGE = {
    "Girls(<18)": "girls", "Women(>=18)": "women", "Female": "female",
    "Boys(<18)": "boys", "Men(>=18)": "men", "Male": "male",
    "Children(<18)": "children", "Adults(>=18)": "adults",
    "Total Beneficiary": "total",
}
CATEGORY = {
    "Host communities": "host_communities", "Refugees": "refugees",
    "Returnees": "returnees", "Internally displaced persons": "idps",
    "Other affected persons": "other",
}


def _people_long(row, code, src):
    out = []

    def add(phase, disagg, grp, v):
        v = _num(v)
        if v is not None:
            out.append({
                "application_code": code, "phase": phase, "disaggregation": disagg,
                "grp": grp, "value": int(v), "source": src,
            })

    for label, grp in SEX_AGE.items():
        add("planned", "sex_age", grp, row.get(f"Planned - {label}"))
        add("planned", "disability", grp, row.get(f"Disabilities - Planned - {label}"))
        add("reached", "disability", grp, row.get(f"Disabilities - Reached - {label}"))
    # reached sex/age columns have inconsistent labels in the GMS export
    reached_map = {
        "Reached Female - Girls(<18)": "girls", "Reached Female - Women(>=18)": "women",
        "Reached - Female": "female", "Reached Male - Boys(<18)": "boys",
        "Reached Male - Men(>=18)": "men", "Reached - Male": "male",
        "Reached - Children(<18)": "children", "Reached - Adults(>=18)": "adults",
        "Reached - Total Beneficiary": "total",
    }
    for label, grp in reached_map.items():
        add("reached", "sex_age", grp, row.get(label))
    for label, grp in CATEGORY.items():
        add("planned", "category", grp, row.get(f"Category - Planned - {label}"))
        add("reached", "category", grp, row.get(f"Category - Reached - {label}"))
    return out


def parse_displacement():
    xl = pd.ExcelFile(F_DISPLACEMENT)
    people, reports = [], []

    src = "yakubu-gms-2020-2024"
    df = _strip_cols(xl.parse("CERF GMS Data 2020-2024"))
    for _, r in df.iterrows():
        code = _s(r.get("Application Code"))
        if not code:
            continue
        people += _people_long(r, code, src)
        reports.append({
            "application_code": code,
            "report_code": _s(r.get("Report Code")),
            "report_focal_point": _s(r.get("Report Focal Point")),
            "language": _s(r.get("Language")),
            "report_deadline": _date(r.get("Report Deadline")),
            "revised_deadline": _date(r.get("Revised Deadline")),
            "cleared": _date(r.get("Cleared")),
            "application_keywords": _s(r.get("Application Keywords")),
            "application_grouping": _s(r.get("Application Grouping")),
            "narr_1a_situation": _s(r.get("1a. Overview of the humanitarian situation")),
            "narr_1b_assistance": _s(r.get("1b. CERF-funded assistance")),
            "narr_2a_situation": _s(r.get("2a. Overview of the humanitarian situation")),
            "narr_2b_assistance": _s(r.get("2b. CERF-funded assistance")),
            "narr_3a_situation": _s(r.get("3a. Overview of the humanitarian situation")),
            "narr_3b_assistance": _s(r.get("3b. CERF-funded assistance provided")),
            "narr_3c_added_value": _s(r.get("3c. CERF’s strategic added value")),
            "source": src,
        })

    src = "yakubu-onegms-2024-2025"
    df = _strip_cols(xl.parse("OneGMS Data 2024-2025"))
    extra = []
    onegms_sex_age = {
        "Girls": "girls", "Boys": "boys", "Women": "women", "Men": "men",
        "Female": "female", "Male": "male", "Children": "children",
        "Adult": "adults", "Total": "total",
    }
    onegms_cat = {
        "Total Host Communities": "host_communities", "Total Refugees": "refugees",
        "Total Returnees": "returnees",
        "Total Internally Displaced People": "idps",
    }
    for _, r in df.iterrows():
        code = _s(r.get("Allocation Code"))
        if not code or not code.startswith("CERF-"):
            continue
        extra.append({
            "application_code": code,
            "is_aa_reported": _bool(r.get("Is AA Allocation")),
            "allocation_keywords": _s(r.get("Allocation Keywords")),
            "is_sudden_onset": _bool(r.get("Is Sudden Onset")),
            "is_slow_onset": _bool(r.get("Is Slow Onset")),
            "response_required_usd": _num(r.get("Total Amount Required For Response")),
            "response_received_usd": _num(r.get("Total Amount Received For Response")),
            "people_affected": (
                int(_num(r.get("Total People Affected By Crisis")))
                if _num(r.get("Total People Affected By Crisis")) is not None else None
            ),
            "source": src,
        })
        for label, grp in onegms_sex_age.items():
            v = _num(r.get(label))
            if v is not None:
                people.append({
                    "application_code": code, "phase": "planned",
                    "disaggregation": "sex_age", "grp": grp, "value": int(v),
                    "source": src,
                })
            dv = _num(r.get(f"Disabled {label}" if label != "Adult" else "Disabled Adult"))
            if dv is not None:
                people.append({
                    "application_code": code, "phase": "planned",
                    "disaggregation": "disability", "grp": grp, "value": int(dv),
                    "source": src,
                })
        for label, grp in onegms_cat.items():
            v = _num(r.get(label))
            if v is not None:
                people.append({
                    "application_code": code, "phase": "planned",
                    "disaggregation": "category", "grp": grp, "value": int(v),
                    "source": src,
                })

    people_df = pd.DataFrame(people).drop_duplicates(
        subset=["application_code", "phase", "disaggregation", "grp", "source"]
    )
    return {
        "cerf_application_people": people_df,
        "cerf_application_report": pd.DataFrame(reports).drop_duplicates(
            subset=["application_code"]
        ),
        "cerf_allocation_extra": pd.DataFrame(extra).drop_duplicates(
            subset=["application_code"]
        ),
    }


# --------------------------------------------------------------------------
# Yakubu: subgrants (full CERF 2020-2025 + curated AA set with localization)
# --------------------------------------------------------------------------

def _subgrant_rows(df, src, is_aa, partner_col, localization=False):
    rows = []
    for _, r in df.iterrows():
        pcode = _s(r.get("Project Code"))
        partner = _s(r.get(partner_col))
        if not pcode or not partner:
            continue
        app = _s(r.get("Application Code"))
        c = iso3_from_application_code(app) or iso3(_s(r.get("Country")))
        rows.append({
            "project_code": pcode, "application_code": app,
            "agency": _s(r.get("Agency")),
            "year": int(_num(r.get("Year"))) if _num(r.get("Year")) is not None else None,
            "window_name": _s(r.get("Window")),
            "country_iso3": c, "country_name": _s(r.get("Country")),
            "emergency_type": _s(r.get("Emergency Type")),
            "project_amount_usd": _num(r.get("Amount Approved")),
            "partner_name": partner,
            "partner_acronym": _s(r.get("Implementing Partner Acronym")),
            "partner_type": _s(r.get("Partner Type")),
            "localization": _s(r.get("Localization")) if localization else None,
            "pre_existing_agreement": _s(r.get("Pre-existing Partnership Agreement")),
            "subgrant_usd": _num(r.get("Subgrants")),
            "is_aa": is_aa,
            "source": src,
        })
    return rows


def parse_subgrants():
    full = _strip_cols(pd.read_excel(F_SUBGRANTS, sheet_name="Sheet3"))
    aa = _strip_cols(pd.read_excel(F_JUN2026, sheet_name="Subgrants - June 2026"))

    aa_rows = _subgrant_rows(aa, "yakubu-subgrants-aa-jun2026", True,
                             "Organizations", localization=True)
    full_rows = _subgrant_rows(full, "yakubu-subgrants-full-2020-2025", False,
                               "Implementing Partner Name")

    aa_df = pd.DataFrame(aa_rows)
    full_df = pd.DataFrame(full_rows)
    # mark AA rows in the full set, and drop those duplicated by the curated set
    aa_apps = set(aa_df["application_code"].dropna())
    full_df["is_aa"] = full_df["application_code"].isin(aa_apps)
    key = ["project_code", "partner_name", "subgrant_usd"]
    aa_keys = set(map(tuple, aa_df[key].fillna(-1).itertuples(index=False)))
    dup = full_df[key].fillna(-1).apply(tuple, axis=1).isin(aa_keys)
    full_df = full_df[~dup]
    out = pd.concat([aa_df, full_df], ignore_index=True)
    out = out.drop_duplicates(subset=["project_code", "partner_name", "subgrant_usd",
                                      "source"])
    return {"cerf_subgrant": out}


# --------------------------------------------------------------------------
# Yakubu: June 2026 workbook (project supplement, CVA, framework-level tables)
# --------------------------------------------------------------------------

def parse_jun2026():
    xl = pd.ExcelFile(F_JUN2026)
    out = {}

    # ---- Agency HQ Report -> cerf_project_supplement
    df = _strip_cols(xl.parse("Agency HQ Report"))
    rows = []
    for _, r in df.iterrows():
        pcode = _s(r.get("ProjectCode"))
        if not pcode:
            continue

        def bi(col, r=r):
            v = _num(r.get(col))
            return int(v) if v is not None else None

        rows.append({
            "project_code": pcode,
            "allocation_code": _s(r.get("AllocationCode")),
            "is_aa": _s(r.get("IS AA?")) == "Yes",
            "gender_marker": _s(r.get("GenderMarker")),
            "gbv_marker": _s(r.get("GBVMarker")),
            "disability_marker": _s(r.get("DisabilityPeopleTargeted")),
            "cash_marker": _s(r.get("CashTransferMarker")),
            "people_receiving_cash": bi("PeopleRecievingCashAssistance"),
            "cva_usd": _num(r.get("CTPAmount")),
            "cva_comments": _s(r.get("CTPComments")),
            "pwd_targeted": bi("PersonsWithDisabilities"),
            "refugees_targeted": bi("Refugees"),
            "returnees_targeted": bi("Returnees"),
            "idps_targeted": bi("IDP"),
            "host_communities_targeted": bi("HostCommunities"),
            "source": "yakubu-agency-hq-jun2026",
        })
    out["cerf_project_supplement"] = pd.DataFrame(rows).drop_duplicates(
        subset=["project_code"]
    )

    # ---- Disbursement Data+CVA -> cerf_cva_history (aggregated to the sheet grain)
    df = _strip_cols(xl.parse("Disbursement Data+CVA June 2026"))
    df = df[df["Country"].notna()].copy()
    df["country_iso3"] = df["Country"].map(iso3)
    for col in ("Amount Approved", "People receiving cash", "Total CVA"):
        df[col] = df[col].map(_num)
    df["Year"] = df["Year"].map(_num)
    df = df[df["Year"].notna()]
    grouped = (
        df.groupby(["country_iso3", "Country", "Agency", "Emergency Type", "Year"],
                   dropna=False)
        .agg(
            amount_approved_usd=("Amount Approved", "sum"),
            people_receiving_cash=("People receiving cash", "sum"),
            cva_usd=("Total CVA", "sum"),
            cva_possible=("CVA Possible", "first"),
            n_source_rows=("CVA Possible", "size"),
        )
        .reset_index()
    )
    out["cerf_cva_history"] = pd.DataFrame({
        "country_iso3": grouped["country_iso3"],
        "country_name": grouped["Country"],
        "agency": grouped["Agency"],
        "emergency_type": grouped["Emergency Type"],
        "year": grouped["Year"].astype("Int64"),
        "amount_approved_usd": grouped["amount_approved_usd"],
        "people_receiving_cash": grouped["people_receiving_cash"].astype("Int64"),
        "cva_usd": grouped["cva_usd"],
        "cva_possible": grouped["cva_possible"],
        "n_source_rows": grouped["n_source_rows"].astype("Int64"),
        "source": "yakubu-cva-jun2026",
    })

    # ---- People covered / Double activations -> people_covered
    covered = []
    df = _strip_cols(xl.parse("People covered"))
    for _, r in df.iterrows():
        c = iso3(_s(r.get("Countries")))
        h, _hr = norm_hazard(_s(r.get("Hazard")))
        if c is None or h is None:
            continue
        v = _num(r.get("People covered"))
        covered.append({
            "country_iso3": c, "hazard": h, "as_of": "2026-06-01",
            "source": "yakubu-people-covered-jun2026",
            "people_covered": int(v) if v is not None else None,
            "remarks": _s(r.get("Remarks")),
        })
    df = xl.parse("Double activations", header=None)
    for i in range(1, len(df)):
        c = iso3(_s(df.iat[i, 0]))
        h, _hr = norm_hazard(_s(df.iat[i, 1]))
        if c is None or h is None:
            continue
        base = _num(df.iat[i, 2])
        double = _s(df.iat[i, 3])
        add = _num(df.iat[i, 4])
        covered.append({
            "country_iso3": c, "hazard": h, "as_of": "2025-12-31",
            "source": "yakubu-double-activations",
            "people_covered": int(base) if base is not None else None,
            "double_activation": double.replace(" ", "_") if double else None,
            "additional_people_covered": int(add) if add is not None else None,
        })
    cov_df = pd.DataFrame(covered).drop_duplicates(
        subset=["country_iso3", "hazard", "as_of", "source"]
    )
    out["people_covered"] = cov_df

    # ---- Co-financing -> prearranged_funding (+ status)
    funding, status = [], []
    df = _strip_cols(xl.parse("Co-financing"))
    for _, r in df.iterrows():
        c = iso3(_s(r.get("Country")))
        h, _hr = norm_hazard(_s(r.get("Hazard")))
        if c is None or h is None:
            continue
        src = "yakubu-cofinancing-jun2026"
        st, st_raw = norm_status(_s(r.get("Status")))
        if st:
            status.append({
                "country_iso3": c, "hazard": h, "as_of": "2026-06-01", "source": src,
                "status": st, "status_raw": st_raw, "comments": _s(r.get("Remarks")),
            })
        for col, kind, fund in (
            ("Pre-arranged (CERF)", "prearranged", "cerf"),
            ("Pre-arranged (Country Fund)", "prearranged", "country_regional"),
            ("Additional co-funding", "cofinancing", "other"),
            ("Non-AA Emergency Funds Mobilised based on forecast",
             "non_aa_mobilised", "other"),
        ):
            amt = _num(r.get(col))
            if amt is not None:
                funding.append({
                    "country_iso3": c, "hazard": h, "year": 2026, "kind": kind,
                    "fund_source": fund, "amount_usd": amt,
                    "remarks": _s(r.get("Remarks")) if kind != "prearranged" else None,
                    "source": src,
                })

    # ---- Pre-arranged (Jun 2026) -> prearranged_funding (+ status)
    df = _strip_cols(xl.parse("Pre-arranged "))
    for _, r in df.iterrows():
        country = _s(r.get("Country"))
        if not country or country == "Total":
            continue
        c = iso3(country)
        h, _hr = norm_hazard(_s(r.get("Shock")))
        if c is None or h is None:
            continue
        src = "yakubu-prearranged-jun2026"
        st, st_raw = norm_status(_s(r.get("Status")))
        if st:
            status.append({
                "country_iso3": c, "hazard": h, "as_of": "2026-06-01", "source": src,
                "status": st, "status_raw": st_raw,
            })
        amt = _num(r.get("Pre-arranged amount"))
        if amt is not None:
            funding.append({
                "country_iso3": c, "hazard": h, "year": 2026, "kind": "prearranged",
                "fund_source": "cerf", "amount_usd": amt, "source": src,
            })

    # ---- New/Extended frameworks 2025 -> prearranged_funding
    df = xl.parse("New-Extended Frameworks(2025)", header=None)
    for i in range(1, len(df)):
        c = iso3(_s(df.iat[i, 11]))
        h, _hr = norm_hazard(_s(df.iat[i, 12]))
        if c is None or h is None or _s(df.iat[i, 11]) == "Total":
            continue
        amt = _num(df.iat[i, 13])
        if amt is not None:
            funding.append({
                "country_iso3": c, "hazard": h, "year": 2025, "kind": "prearranged",
                "fund_source": "cerf", "amount_usd": amt,
                "funding_change": (_s(df.iat[i, 14]) or "").lower() or None,
                "source": "yakubu-new-extended-2025",
            })

    # ---- Insurance portfolio (Jan 2026 snapshot) -> status + funding
    df = _strip_cols(xl.parse("A) 2026 Portfolio for Insurance", header=1))
    for _, r in df.iterrows():
        country = _s(r.get("Country"))
        if not country or str(country).startswith(("Total", "Row 30")):
            continue
        c = iso3(country)
        h, _hr = norm_hazard(_s(r.get("Hazard")))
        if c is None or h is None:
            continue
        src = "yakubu-insurance-2026"
        st, st_raw = norm_status(_s(r.get("Jan 2026 Status")))
        exp, _exp_raw = norm_status(_s(r.get("2026 Status expected for calculation purposes")))
        if st:
            status.append({
                "country_iso3": c, "hazard": h, "as_of": "2026-01-15", "source": src,
                "status": st, "status_raw": st_raw, "expected_status": exp,
                "q1_ready": _bool(r.get("Q1 ready")),
                "comments": _s(r.get("Comments ")),
            })
        cols = {str(col).strip(): col for col in df.columns}
        amt = _num(r.get(cols.get("$ Pre-arranged")))
        if amt is not None:
            funding.append({
                "country_iso3": c, "hazard": h, "year": 2026, "kind": "prearranged",
                "fund_source": "cerf", "amount_usd": amt, "source": src,
            })

    out["prearranged_funding"] = pd.DataFrame(funding).drop_duplicates(
        subset=["country_iso3", "hazard", "year", "kind", "fund_source", "source"]
    )
    out["framework_status"] = pd.DataFrame(status).drop_duplicates(
        subset=["country_iso3", "hazard", "as_of", "source"]
    )

    # ---- Sector Data - Pre-arranged -> prearranged_sector_budget
    df = _strip_cols(xl.parse("Sector Data - Pre-arranged"))
    rows = []
    for _, r in df.iterrows():
        country = _s(r.get("Country"))
        agency = _s(r.get("Agency"))
        sector = _s(r.get("Sector"))
        if not country or not agency or not sector:
            continue
        c = iso3(country)
        h, _hr = norm_hazard(_s(r.get("Shock")))
        if c is None or h is None:
            continue
        rows.append({
            "country_iso3": c, "hazard": h, "window_name": subunit_of(country),
            "agency": agency, "sector": sector,
            "amount_usd": _num(r.get("Amount")),
            "year_label": _s(r.get("Year")),
            "status": _s(r.get("Status")),
            "source": "yakubu-sector-prearranged-jun2026",
        })
    sect = pd.DataFrame(rows)
    if not sect.empty:
        sect = (
            sect.groupby(
                ["country_iso3", "hazard", "window_name", "agency", "sector", "year_label"],
                dropna=False, as_index=False,
            )
            .agg(amount_usd=("amount_usd", "sum"), status=("status", "first"),
                 source=("source", "first"))
        )
    out["prearranged_sector_budget"] = sect

    return out


ALL_PARSERS = [
    parse_planning,
    parse_reporting,
    parse_activations,
    parse_alloc_analysis,
    parse_displacement,
    parse_subgrants,
    parse_jun2026,
]


def parse_all():
    """Run every parser and concatenate outputs per table."""
    merged: dict[str, list] = {}
    for fn in ALL_PARSERS:
        for table, df in fn().items():
            if df is not None and not df.empty:
                merged.setdefault(table, []).append(df)
    out = {t: pd.concat(dfs, ignore_index=True) for t, dfs in merged.items()}
    if UNMAPPED:
        print(f"WARNING unmapped countries (dropped): {sorted(UNMAPPED)}")
    return out
